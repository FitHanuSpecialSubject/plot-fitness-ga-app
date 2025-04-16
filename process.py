import os
import pandas as pd
import matplotlib.pyplot as plt
import openpyxl
from pathlib import Path


script_path = Path(__file__).resolve().parent
source_folder = os.path.join(script_path, "source-excel-files")
result_folder = os.path.join(script_path, "result-images")

os.makedirs(result_folder, exist_ok=True)
xlsx_files = [f for f in os.listdir(source_folder) if f.endswith('.xlsx')]

if not xlsx_files:
    raise FileNotFoundError(f"No .xlsx file found in the \n{source_folder} folder. \n Place one in.")

source_xlsx_file = os.path.join(source_folder, xlsx_files[0])


# Load workbook
wb = openpyxl.load_workbook(source_xlsx_file)
source_sheet = wb["Sheet 1"]

num_iteration = int(source_sheet['A'][-1].value)
row_length = 0
for cell in source_sheet[1]:
    if cell.value is None:
        break
    row_length += 1
num_algorithms = row_length - 1



data = {}
data['Iteration'] = [i for i in range(1, num_iteration + 1)]
for i in range(1, num_algorithms + 1):
    algorithm_name = source_sheet.cell(row=1, column=i + 1).value
    print(algorithm_name)
    data[algorithm_name] = []
    for j in range(0, num_iteration):
        data[algorithm_name].append(source_sheet.cell(row=j + 2, column=i + 1).value)


for k, v in data.items():
    print(k, v)

df = pd.DataFrame(data)

# Plotting
plt.figure(figsize=(10, 6))
for column in df.columns[1:]:
    plt.plot(df['Iteration'], df[column], label=column, marker='o')

plt.title('Fitness Values Comparison Across Algorithms')
plt.xlabel('Iteration')
plt.ylabel('Fitness Value')
plt.legend()
plt.grid(True)
plt.tight_layout()


output_file = os.path.join(result_folder, "fitness_comparison.png")
plt.savefig(output_file)
print(f"Plot saved to {output_file}")
